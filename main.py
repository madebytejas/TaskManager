import time
from plyer import notification
import openpyxl
import datetime




def ReadData(i):
    global taskname,tasktime,taskdetails
    TaskName = dataframe1.cell(row = i, column = 2)
    TaskTime = dataframe1.cell(row = i, column = 5)
    TaskDetails = dataframe1.cell(row = i, column = 4)
    taskname = TaskName.value
    print (taskname)
    tasktime = TaskTime.value
    taskdetails = TaskDetails.value
    if (taskname==None):
        exit(1)

def  sleepTime(Task_Time):
    global sleeptime,CrTimeMin
    CrTime = datetime.datetime.now()
    NowHour = CrTime.hour*60
    NowMin = CrTime.minute
    CrTimeMin=NowHour+NowMin
    # print(f"{CrTime}={CrTimeMin}")
    # Task_Time = int(Task_Time)
    sleeptime = Task_Time-CrTimeMin
    print("sleeptime="+str(sleeptime))
    if sleeptime>0 :
        time.sleep((sleeptime*60)-15)

def taskReminder(Task_name,Task_Details):

    notification.notify(
        title = f"**Gentle Reminder {Task_name} **",
        message = f"{Task_Details}",
        app_icon =r'D:\vscode projects\Task Shedular and remider\bell.ico',
        timeout=5
    )

path = "D:\\vscode projects\\Task Shedular and remider\\Tasks.xlsx"
dataframe = openpyxl.load_workbook(path, data_only=True)
dataframe1 = dataframe.active
Mrows = dataframe1.max_row

for i in range(2,Mrows):
    ReadData(i)
    sleepTime(tasktime)
    if sleeptime>0:
        taskReminder(taskname,taskdetails)
    
    